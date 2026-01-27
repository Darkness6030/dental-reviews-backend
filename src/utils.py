from io import BytesIO
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def export_rows_to_excel(
    rows_data: List[Dict],
    max_column_width: int = 40,
    width_padding_ratio: float = 0.15
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active

    column_names = list(rows_data[0].keys())
    header_style = Font(bold=True)

    for column_index, column_name in enumerate(column_names, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.font = header_style

    for row_index, row in enumerate(rows_data, start=2):
        for column_index, column_name in enumerate(column_names, start=1):
            worksheet.cell(row=row_index, column=column_index, value=row.get(column_name))

    for column_index, column_name in enumerate(column_names, start=1):
        column_letter = get_column_letter(column_index)
        max_text_length = len(str(column_name))

        for cell in worksheet[column_letter]:
            if cell.value is not None:
                max_text_length = max(max_text_length, len(str(cell.value)))

        padded_width = max_text_length * (1 + width_padding_ratio)
        worksheet.column_dimensions[column_letter].width = min(
            padded_width,
            max_column_width
        )

    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()
